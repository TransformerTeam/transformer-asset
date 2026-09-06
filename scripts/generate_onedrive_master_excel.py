import json
import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
plan_json_path = os.path.join(base_dir, 'temp_initial_tasks.json')
csv_path = os.path.join(base_dir, 'HealthIndexSum.csv')
out_path = os.path.join(base_dir, 'GPSC_Transformer_Asset_Master.xlsx')

wb = openpyxl.Workbook()

# Sheet 1: Transformer_Plan
ws_plan = wb.active
ws_plan.title = 'Transformer_Plan'

headers_plan = [
    'WBS', 'SAP_MO', 'Parent_WBS', 'Task_Description', 'Lead',
    'Transformer_Tag', 'Plant', 'Category', 'Start_Date', 'End_Date',
    'Days', 'Progress', 'Cost'
]

header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

ws_plan.append(headers_plan)
for col_num, h in enumerate(headers_plan, 1):
    cell = ws_plan.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

if os.path.exists(plan_json_path):
    with open(plan_json_path, 'r', encoding='utf-8') as f:
        tasks = json.load(f)
else:
    tasks = []

for r_idx, t in enumerate(tasks, 2):
    row_data = [
        t.get('wbs', ''),
        str(t.get('sapMo', '')),
        t.get('parentId') or '',
        t.get('task', ''),
        t.get('lead', ''),
        t.get('tr', ''),
        t.get('plant', ''),
        t.get('cat', ''),
        t.get('start', ''),
        t.get('end', ''),
        t.get('days', 1),
        t.get('progress', 0),
        t.get('cost', 0)
    ]
    ws_plan.append(row_data)
    for col_idx in range(1, len(row_data) + 1):
        c = ws_plan.cell(row=r_idx, column=col_idx)
        c.border = thin_border
        c.font = Font(name='Segoe UI', size=10)
        if col_idx in [1, 2, 3, 5, 6, 7, 8, 9, 10]:
            c.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx in [11, 12, 13]:
            c.alignment = Alignment(horizontal='right', vertical='center')
        else:
            c.alignment = Alignment(horizontal='left', vertical='center')

for col in ws_plan.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_plan.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Sheet 2: Health_Index_Fleet
ws_fleet = wb.create_sheet(title='Health_Index_Fleet')
fleet_rows = []
if os.path.exists(csv_path):
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        fleet_rows = list(reader)

if fleet_rows:
    for r_idx, row in enumerate(fleet_rows, 1):
        ws_fleet.append(row)
        for col_idx in range(1, len(row) + 1):
            c = ws_fleet.cell(row=r_idx, column=col_idx)
            c.border = thin_border
            if r_idx == 1:
                c.fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
                c.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.font = Font(name='Segoe UI', size=9)

for col in ws_fleet.columns:
    max_len = max(len(str(cell.value or '')) for cell in col[:20])
    col_letter = get_column_letter(col[0].column)
    ws_fleet.column_dimensions[col_letter].width = max(min(max_len + 3, 30), 10)

# Sheet 3: System_Guide
ws_info = wb.create_sheet(title='System_Guide')
ws_info.append(['GPSC Transformer Asset Management - Enterprise OneDrive Master Dataset'])
ws_info.append([])
ws_info.append(['How to use this Excel workbook:'])
ws_info.append(['1. Save this file into your team shared OneDrive or Microsoft Teams / SharePoint folder.'])
ws_info.append(['2. Sheet "Transformer_Plan" contains the 52-week outage & maintenance tasks.'])
ws_info.append(['3. Sheet "Health_Index_Fleet" contains the latest fleet diagnostic evaluation and health indices.'])
ws_info.append(['4. You can edit tasks directly here in Excel Online, or edit via the web dashboard.'])
ws_info.append(['5. When changes are saved, the web application will automatically sync and update for all team members.'])

ws_info['A1'].font = Font(name='Segoe UI', size=14, bold=True, color='1E293B')
for r in range(3, 9):
    ws_info[f'A{r}'].font = Font(name='Segoe UI', size=11)
ws_info.column_dimensions['A'].width = 90

wb.save(out_path)
print(f'Successfully generated: {out_path}')

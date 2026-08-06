import os
import sys
import io
import csv
from datetime import datetime
import openpyxl

# Set output encoding to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\NB\OneDrive - GPSC\GPSC Group Transformer Assessment\Transformer Asset Managment GPSC GROUP Rev.25.xlsm"
OUTPUT_DIR = r"C:\Users\NB\Downloads\TR Asset\TestData"
HEALTH_SUM_PATH = r"C:\Users\NB\Downloads\TR Asset\HealthIndexSum.csv"

# Target sheets to export
SHEET_NAMES = [
    'SAPorder',
    'FactoryData',
    'VisualData',
    'WindingPFData',
    'IRandPIData',
    'BushingPFData',
    'BushingInfo',
    'SurgePFData',
    'SurgeInfo',
    'ExcitingData',
    'RatioData',
    'WindingData',
    'SingleShortData',
    'ThreeShortData',
    'FRAData',
    'DFRData',
    'DRMData',
    'PDonlineData',
    'ThermoScanData',
    'TRinfo2',
    'TRHistory',
    'MTOilData',
    'OLTCOilData'
]

def export_sheet(ws, out_filepath):
    rows = []
    for r in ws.iter_rows(values_only=True):
        row_vals = []
        has_val = False
        for cell in r:
            if cell is None:
                row_vals.append("")
            elif isinstance(cell, datetime):
                row_vals.append(cell.strftime("%Y-%m-%d %H:%M:%S") if cell.time() else cell.strftime("%Y-%m-%d"))
                has_val = True
            else:
                s_val = str(cell).strip()
                if s_val != "":
                    has_val = True
                row_vals.append(s_val)
        if has_val:
            rows.append(row_vals)
    
    if rows:
        # Trim trailing empty columns across all rows if needed, or write directly
        os.makedirs(os.path.dirname(out_filepath), exist_ok=True)
        with open(out_filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        print(f"Exported {ws.title} -> {out_filepath} ({len(rows)} rows)")
    else:
        print(f"Sheet {ws.title} is empty or has no data rows.")

def export_health_index_sum(ws, out_filepath):
    headers = []
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        headers = list(row)

    clean_headers = [str(h).replace('\n', ' ').strip() if h is not None else '' for h in headers[:56]]

    data_rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        row_data = list(row)[:56]
        non_empty = [v for v in row_data if v is not None and str(v).strip() != '']
        if not non_empty:
            continue
        eq_name = row_data[1] if len(row_data) > 1 else None
        serial = row_data[2] if len(row_data) > 2 else None
        if not eq_name and not serial:
            continue
        
        cleaned = []
        for val in row_data:
            if isinstance(val, datetime):
                cleaned.append(val.strftime('%m/%d/%Y'))
            elif val is None:
                cleaned.append('')
            else:
                cleaned.append(str(val).strip())
        data_rows.append(cleaned)

    with open(out_filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(clean_headers)
        writer.writerows(data_rows)
    print(f"Exported HealthIndexSum -> {out_filepath} ({len(data_rows)} rows)")

def main():
    print(f"Starting export from {EXCEL_PATH} at {datetime.now()}")
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file does not exist at {EXCEL_PATH}")
        sys.exit(1)

    import shutil
    temp_path = "temp_export.xlsm"
    try:
        shutil.copy2(EXCEL_PATH, temp_path)
    except Exception as e:
        print(f"Failed to copy file to bypass lock: {e}")
        temp_path = EXCEL_PATH

    try:
        wb = openpyxl.load_workbook(temp_path, read_only=True, keep_vba=False, data_only=True)
    except Exception as e:
        print(f"Failed to load workbook: {e}")
        if os.path.exists(temp_path) and temp_path != EXCEL_PATH:
            os.remove(temp_path)
        sys.exit(1)
    all_sheets = wb.sheetnames

    # Export HealthIndexSum
    if 'HealthIndexSum' in all_sheets:
        export_health_index_sum(wb['HealthIndexSum'], HEALTH_SUM_PATH)

    # Export target TestData sheets
    for s_name in SHEET_NAMES:
        if s_name in all_sheets:
            out_file = os.path.join(OUTPUT_DIR, f"{s_name}.csv")
            export_sheet(wb[s_name], out_file)
        else:
            print(f"WARNING: Sheet '{s_name}' not found in workbook.")

    # Auto-evaluate missing Health Index scores & sync health_data.js
    try:
        from evaluate_all_health_index import main as eval_main
        eval_main()
    except Exception as e:
        print(f"Error evaluating Health Index: {e}")

    # Close workbook and clean up temp copy
    try:
        wb.close()
    except Exception:
        pass
    if os.path.exists(temp_path) and temp_path != EXCEL_PATH:
        try:
            os.remove(temp_path)
        except Exception:
            pass

    print(f"All exports and evaluations completed successfully at {datetime.now()}.")

if __name__ == "__main__":
    main()

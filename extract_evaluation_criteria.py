import openpyxl
import json
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_PATH = r'C:\Users\NB\OneDrive - GPSC\GPSC Group Transformer Assessment\Transformer Asset Managment GPSC GROUP Rev.25.xlsm'
OUT_JSON = r'C:\Users\NB\Downloads\TR Asset\evaluation_criteria.json'
OUT_JS = r'C:\Users\NB\Downloads\TR Asset\evaluation_criteria.js'

print(f"Loading workbook {EXCEL_PATH}...")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, keep_vba=False, data_only=True)

if 'Evaluation' not in wb.sheetnames:
    print("Evaluation sheet not found!")
    sys.exit(1)

ws = wb['Evaluation']
items = []

category = "General Parts (Weight Score 5%)"
for r in range(6, ws.max_row + 1):
    c1 = ws.cell(row=r, column=1).value
    c10 = ws.cell(row=r, column=10).value
    c16 = ws.cell(row=r, column=16).value
    c18 = ws.cell(row=r, column=18).value

    s1 = str(c1).strip() if c1 else ""
    s10 = str(c10).strip() if c10 else ""
    s16 = str(c16).strip() if c16 else ""
    s18 = str(c18).strip() if c18 else ""

    # Category header detection
    if s1 and not s10 and not s16 and not s18:
        if s1 in ['General', 'General Parts', 'General Parts (Weight Score 5%)']:
            category = "General Parts (Weight Score 5%)"
            continue
        elif s1 in ['Magnetic Core', 'High Voltage Winding', 'Low Voltage Winding', 'Tertiary Winding',
                  'Insulating Oil in Main Tank', 'Insulating Oil in OLTC', 'Surge Arrester', 'Bushing',
                  'OLTC', 'Visual Inspection', 'Grounding Measurement and Test', 'Neutral Ground Resistor',
                  'Cooling System Inspection', 'General Health Index', 'Impact Index', 'Summary']:
            category = s1
            continue

    if s1 or s10 or s16 or s18:
        item_title = s1
        if "Core-Ground" in s1:
            item_title = "Core-Ground Insulation Resistance"
        elif "Clamp-Ground" in s1:
            item_title = "Clamp-Ground Insulation Resistance"
        elif "Core-Clamp" in s1:
            item_title = "Core-Clamp Insulation Resistance"

        items.append({
            "row": r,
            "category": category,
            "item": item_title,
            "criteria": s10,
            "standard": s16,
            "recommendation": s18
        })

# Explicitly add General Parts (Weight Score 5%) items requested by user
general_parts_items = [
    {
        "row": 701,
        "category": "General Parts (Weight Score 5%)",
        "item": "Visual Inspection",
        "criteria": "-",
        "standard": "GPSC Criteria",
        "recommendation": "Score: 5/5 (Normal)"
    },
    {
        "row": 702,
        "category": "General Parts (Weight Score 5%)",
        "item": "Grounding Resistance",
        "criteria": "< 1 Ω",
        "standard": "GPSC Criteria",
        "recommendation": "Score: 5/5 (Normal)"
    },
    {
        "row": 703,
        "category": "General Parts (Weight Score 5%)",
        "item": "Contact Resistance",
        "criteria": "≤ 10 µΩ",
        "standard": "GPSC Criteria",
        "recommendation": "Score: 5/5 (Normal)"
    },
    {
        "row": 704,
        "category": "General Parts (Weight Score 5%)",
        "item": "Neutral Ground Inspection",
        "criteria": "-",
        "standard": "IEC 60076-25:2023",
        "recommendation": "Score: 5/5 (Normal)"
    }
]
items = general_parts_items + [x for x in items if not (x.get('item') == 'Criteria' and x.get('category') == 'General Parts (Weight Score 5%)')]

print(f"Extracted {len(items)} evaluation criteria items.")

with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(items, f, indent=2, ensure_ascii=False)

js_content = f"const EVALUATION_CRITERIA_DATA = {json.dumps(items, indent=2, ensure_ascii=False)};\n"
with open(OUT_JS, 'w', encoding='utf-8') as f:
    f.write(js_content)

print(f"Exported to {OUT_JS} successfully.")

import openpyxl
import json
import csv
import os

excel_path = r'C:\Users\NB\OneDrive - GPSC\1. Database\GPSC GROUP TRANSFORMER LIST_OLTC.xlsx'
if not os.path.exists(excel_path):
    print(f"File not found: {excel_path}")
    exit(1)

wb = openpyxl.load_workbook(excel_path, data_only=True)

# 1. Sheet 1: GPSC RROUP TRANSFORMER_OLTC
sh1 = wb['GPSC RROUP TRANSFORMER_OLTC']
headers1 = [sh1.cell(2, c).value for c in range(1, sh1.max_column + 1)]

records = []
for r in range(3, sh1.max_row + 1):
    row_vals = [sh1.cell(r, c).value for c in range(1, sh1.max_column + 1)]
    if not any(row_vals):
        continue
    rec = {}
    for h, v in zip(headers1, row_vals):
        if h and v is not None:
            rec[str(h).strip()] = str(v).strip()
    
    parent_sn = rec.get('Parent Serial No.') or ''
    kks = rec.get('KKS No.') or ''
    equip_name = rec.get('Equip. Name') or ''
    
    if parent_sn or kks or equip_name:
        schedule = {}
        for y in range(2016, 2033):
            ystr = str(y)
            if ystr in rec:
                schedule[ystr] = rec[ystr]
        
        entry = {
            'item': rec.get('Item', ''),
            'plantSite': rec.get('Pant Site') or rec.get('Site', ''),
            'location': rec.get('Location', ''),
            'kksNo': kks,
            'equipmentName': equip_name,
            'parentSerialNo': parent_sn,
            'vectorGroup': rec.get('Vector Group', ''),
            'serviceType': rec.get('Service Type', ''),
            'oltcManufacturer': rec.get('Manufacture', ''),
            'oltcModelYear': rec.get('Model Year', ''),
            'oltcModelType': rec.get('Model Type', ''),
            'oltcAge': rec.get('Age', ''),
            'oltcSerialNo': rec.get('Serail No.', ''),
            'motorDrive': rec.get('Motor Drive', ''),
            'counterOperated': rec.get('Counter Operated', ''),
            'oilFilterType': rec.get('Oil filter Type', ''),
            'tapNo': rec.get('Tap No.', ''),
            'maintenanceTime': rec.get('Maintenance Time', ''),
            'contactLife': rec.get('Contact life', ''),
            'transitionResistorOhm': rec.get('Transistion Resistor (Ohm)', ''),
            'stepVoltageV': rec.get('Stepvoltage (V)', ''),
            'currentA': rec.get('current (A)', ''),
            'description': rec.get('Decription', ''),
            'lastInspection': rec.get('Last Inspection', ''),
            'nextDue': rec.get('Next Due', ''),
            'remark': rec.get('Remark', ''),
            'schedule': schedule
        }
        records.append(entry)

print(f"Extracted {len(records)} OLTC records from Sheet 1.")

# Write JS file
js_path = r'C:\Users\NB\Downloads\TR Asset\oltc_data.js'
with open(js_path, 'w', encoding='utf-8') as f:
    f.write('// Auto-generated OLTC Data from GPSC GROUP TRANSFORMER LIST_OLTC.xlsx\n')
    f.write('const OLTC_DATA = ')
    json.dump(records, f, indent=2, ensure_ascii=False)
    f.write(';\n\n')
    f.write('if (typeof module !== "undefined") { module.exports = OLTC_DATA; }\n')

print(f"Successfully generated {js_path}")

# Write CSV file
csv_path = r'C:\Users\NB\Downloads\TR Asset\TestData\OLTCData.csv'
if records:
    fieldnames = ['parentSerialNo', 'kksNo', 'equipmentName', 'plantSite', 'location', 'vectorGroup', 'serviceType', 'oltcManufacturer', 'oltcModelYear', 'oltcModelType', 'oltcSerialNo', 'motorDrive', 'counterOperated', 'tapNo', 'transitionResistorOhm', 'maintenanceTime', 'contactLife', 'stepVoltageV', 'currentA', 'lastInspection', 'nextDue', 'description', 'remark']
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        for r in records:
            writer.writerow(r)

print(f"Successfully generated {csv_path}")
